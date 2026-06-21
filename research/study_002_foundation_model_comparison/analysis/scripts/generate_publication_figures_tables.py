#!/usr/bin/env python3
"""Generate publication-ready Study 002 figures and tables.

The script is dependency-light: it uses only openpyxl and Pillow, both already
available in the project environment. It regenerates public-safe manuscript
figures and summary tables from the analysis-ready Study 002 workbook and the
operational-efficiency/task-stratified outputs.

Outputs:
  results/figures_publication/*.png
  results/publication_tables_v2.xlsx
  results/publication_tables_v2.md
"""

from __future__ import annotations

from collections import defaultdict
from math import sqrt
from pathlib import Path
from statistics import mean, median
from typing import Any, Dict, Iterable, List, Sequence, Tuple

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from PIL import Image, ImageDraw, ImageFont

ROOT = Path(__file__).resolve().parents[1]
STUDY_ROOT = ROOT.parent
DATASET = STUDY_ROOT / "datasets" / "merged" / "Agentic_AI_Experiments_Main_Study002_V1.4.4_270Runs_AnalysisReady.xlsx"
EFFICIENCY = STUDY_ROOT / "analysis" / "statistical_outputs" / "operational_efficiency.xlsx"
TASK_STRATIFIED = STUDY_ROOT / "analysis" / "statistical_outputs" / "task_stratified_analysis.xlsx"
OUT_DIR = STUDY_ROOT / "results"
FIG_DIR = OUT_DIR / "figures_publication"
TABLE_XLSX = OUT_DIR / "publication_tables_v2.xlsx"
TABLE_MD = OUT_DIR / "publication_tables_v2.md"

DATA_SHEET = "Agentic_AI_Experiments_V1.0"
PROVIDER_COL = "Provider"
WORKFLOW_COL = "workflow_type"
CATEGORY_COL = "task_category_final"
DIFFICULTY_COL = "task_difficulty_final"
QUALITY_COL = "quality_score"
CONFIDENCE_COL = "confidence"
COST_COL = "cost_usd"
DURATION_COL = "duration_sec"
TOKENS_COL = "total_tokens"
CALLS_COL = "llm_call_count"

PROVIDER_ORDER = ["OpenAI", "Google", "Anthropic"]
WORKFLOW_ORDER = ["basic_agent", "planner_executor", "planner_executor_reviewer"]
CATEGORY_ORDER = ["Knowledge", "Reasoning", "Coding"]
DIFFICULTY_ORDER = ["easy", "medium", "hard"]

BG = "white"
INK = "#1E293B"
MUTED = "#64748B"
GRID = "#CBD5E1"
BLUE = "#2563EB"
GREEN = "#16A34A"
ORANGE = "#EA580C"
PURPLE = "#7C3AED"
RED = "#DC2626"


def font(size: int, bold: bool = False) -> ImageFont.FreeTypeFont | ImageFont.ImageFont:
    candidates = [
        "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf" if bold else "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
        "/usr/share/fonts/truetype/liberation2/LiberationSans-Bold.ttf" if bold else "/usr/share/fonts/truetype/liberation2/LiberationSans-Regular.ttf",
    ]
    for candidate in candidates:
        if Path(candidate).exists():
            return ImageFont.truetype(candidate, size)
    return ImageFont.load_default()


def as_float(value: Any) -> float | None:
    if value is None:
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def read_records(path: Path, sheet_name: str) -> List[Dict[str, Any]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet {sheet_name!r} not found in {path}. Available: {wb.sheetnames}")
    ws = wb[sheet_name]
    headers = [str(h).strip() if h is not None else "" for h in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        record = dict(zip(headers, row))
        if any(value is not None for value in record.values()):
            rows.append(record)
    return rows


def values_for(rows: List[Dict[str, Any]], metric: str) -> List[float]:
    return [v for row in rows if (v := as_float(row.get(metric))) is not None]


def stddev(values: Sequence[float]) -> float | None:
    vals = list(values)
    if len(vals) < 2:
        return None
    m = mean(vals)
    return sqrt(sum((x - m) ** 2 for x in vals) / (len(vals) - 1))


def summarize(rows: List[Dict[str, Any]], group_cols: Tuple[str, ...]) -> List[Dict[str, Any]]:
    groups: Dict[Tuple[Any, ...], List[Dict[str, Any]]] = defaultdict(list)
    for row in rows:
        groups[tuple(row.get(col) for col in group_cols)].append(row)
    out = []
    for key in sorted(groups, key=lambda parts: tuple(str(part) for part in parts)):
        group = groups[key]
        rec = {col: value for col, value in zip(group_cols, key)}
        rec["N"] = len(group)
        for metric in (QUALITY_COL, CONFIDENCE_COL, COST_COL, DURATION_COL, TOKENS_COL, CALLS_COL):
            vals = values_for(group, metric)
            rec[f"mean_{metric}"] = mean(vals) if vals else None
            rec[f"median_{metric}"] = median(vals) if vals else None
            rec[f"sd_{metric}"] = stddev(vals)
        out.append(rec)
    return out


def fmt(value: Any, digits: int = 3) -> str:
    if value is None:
        return "—"
    if isinstance(value, float):
        return f"{value:.{digits}f}"
    return str(value)


def draw_title(draw: ImageDraw.ImageDraw, title: str, subtitle: str, width: int) -> None:
    draw.text((50, 30), title, fill=INK, font=font(32, True))
    draw.text((50, 72), subtitle, fill=MUTED, font=font(18))
    draw.line((50, 108, width - 50, 108), fill=GRID, width=2)


def color_scale(value: float, min_v: float, max_v: float) -> Tuple[int, int, int]:
    if max_v <= min_v:
        t = 0.5
    else:
        t = (value - min_v) / (max_v - min_v)
    # Light blue to dark blue.
    start = (219, 234, 254)
    end = (29, 78, 216)
    return (
        int(start[0] + (end[0] - start[0]) * t),
        int(start[1] + (end[1] - start[1]) * t),
        int(start[2] + (end[2] - start[2]) * t),
    )


def text_center(draw: ImageDraw.ImageDraw, box: Tuple[int, int, int, int], text: str, fill: str | Tuple[int, int, int], size: int = 18, bold: bool = False) -> None:
    f = font(size, bold)
    bbox = draw.textbbox((0, 0), text, font=f)
    tw, th = bbox[2] - bbox[0], bbox[3] - bbox[1]
    x1, y1, x2, y2 = box
    draw.text((x1 + (x2 - x1 - tw) / 2, y1 + (y2 - y1 - th) / 2), text, fill=fill, font=f)


def heatmap_provider_workflow(pw_rows: List[Dict[str, Any]]) -> None:
    values = {(r[PROVIDER_COL], r[WORKFLOW_COL]): as_float(r.get(f"mean_{QUALITY_COL}")) for r in pw_rows}
    vals = [v for v in values.values() if v is not None]
    min_v, max_v = min(vals), max(vals)
    width, height = 1400, 820
    im = Image.new("RGB", (width, height), BG)
    d = ImageDraw.Draw(im)
    draw_title(d, "Figure 1. Mean operational quality proxy by provider and workflow", "Values are workflow-generated quality_score means; interpret as operational proxy, not human rating.", width)
    left, top = 270, 190
    cell_w, cell_h = 310, 145
    for j, workflow in enumerate(WORKFLOW_ORDER):
        text_center(d, (left + j * cell_w, top - 70, left + (j + 1) * cell_w, top - 10), workflow.replace("_", "\n"), INK, 19, True)
    for i, provider in enumerate(PROVIDER_ORDER):
        text_center(d, (55, top + i * cell_h, left - 25, top + (i + 1) * cell_h), provider, INK, 22, True)
        for j, workflow in enumerate(WORKFLOW_ORDER):
            x1, y1 = left + j * cell_w, top + i * cell_h
            x2, y2 = x1 + cell_w - 12, y1 + cell_h - 12
            v = values.get((provider, workflow))
            fill = color_scale(v or 0, min_v, max_v)
            d.rounded_rectangle((x1, y1, x2, y2), radius=18, fill=fill, outline="white", width=4)
            text_color = "white" if v is not None and v > (min_v + max_v) / 2 else INK
            text_center(d, (x1, y1, x2, y2 - 18), fmt(v, 3), text_color, 30, True)
            text_center(d, (x1, y2 - 35, x2, y2 - 5), "mean quality", text_color, 14)
    d.text((50, height - 65), "Source: Study 002 analysis-ready dataset, n=270. Each provider × workflow cell has 30 runs.", fill=MUTED, font=font(16))
    im.save(FIG_DIR / "figure_01_quality_heatmap_provider_workflow.png", quality=95)


def grouped_bar_by_category(category_rows: List[Dict[str, Any]]) -> None:
    data = {r[CATEGORY_COL]: as_float(r.get(f"mean_{QUALITY_COL}")) for r in category_rows}
    width, height = 1200, 760
    im = Image.new("RGB", (width, height), BG)
    d = ImageDraw.Draw(im)
    draw_title(d, "Figure 2. Mean quality proxy by task category", "Task bank is category-balanced: 10 Knowledge, 10 Reasoning, 10 Coding tasks.", width)
    x0, y0, chart_w, chart_h = 160, 180, 900, 430
    d.line((x0, y0, x0, y0 + chart_h), fill=INK, width=2)
    d.line((x0, y0 + chart_h, x0 + chart_w, y0 + chart_h), fill=INK, width=2)
    for t in [0.0, 0.25, 0.5, 0.75, 1.0]:
        y = y0 + chart_h - t * chart_h
        d.line((x0 - 8, y, x0 + chart_w, y), fill=GRID if t else INK, width=1)
        d.text((65, y - 11), f"{t:.2f}", fill=MUTED, font=font(15))
    colors = [BLUE, GREEN, ORANGE]
    bar_w = 150
    gap = 140
    for i, category in enumerate(CATEGORY_ORDER):
        v = data.get(category) or 0
        x = x0 + 95 + i * (bar_w + gap)
        y = int(y0 + chart_h - v * chart_h)
        d.rounded_rectangle((x, y, x + bar_w, y0 + chart_h), radius=12, fill=colors[i])
        text_center(d, (x - 40, y - 42, x + bar_w + 40, y - 5), fmt(v, 3), INK, 20, True)
        text_center(d, (x - 80, y0 + chart_h + 18, x + bar_w + 80, y0 + chart_h + 55), category, INK, 19, True)
    d.text((50, height - 70), "Difficulty labels are secondary annotations and were not the primary balancing criterion.", fill=MUTED, font=font(16))
    im.save(FIG_DIR / "figure_02_quality_by_task_category.png", quality=95)


def cost_quality_tradeoff(pw_rows: List[Dict[str, Any]]) -> None:
    width, height = 1300, 850
    im = Image.new("RGB", (width, height), BG)
    d = ImageDraw.Draw(im)
    draw_title(d, "Figure 3. Cost–quality trade-off by provider × workflow", "Higher is better on y-axis; lower cost is better on x-axis. Bubble labels show provider/workflow.", width)
    x0, y0, chart_w, chart_h = 130, 170, 1000, 500
    points = []
    for r in pw_rows:
        cost = as_float(r.get(f"mean_{COST_COL}"))
        quality = as_float(r.get(f"mean_{QUALITY_COL}"))
        tokens = as_float(r.get(f"mean_{TOKENS_COL}"))
        if cost is not None and quality is not None:
            points.append((cost, quality, tokens or 0, str(r[PROVIDER_COL]), str(r[WORKFLOW_COL])))
    min_cost, max_cost = min(p[0] for p in points), max(p[0] for p in points)
    min_q, max_q = min(p[1] for p in points), max(p[1] for p in points)
    q_pad = 0.04
    d.rectangle((x0, y0, x0 + chart_w, y0 + chart_h), outline=INK, width=2)
    for t in range(6):
        frac = t / 5
        x = x0 + frac * chart_w
        y = y0 + chart_h - frac * chart_h
        d.line((x, y0, x, y0 + chart_h), fill=GRID, width=1)
        d.line((x0, y, x0 + chart_w, y), fill=GRID, width=1)
    colors = {"OpenAI": BLUE, "Google": GREEN, "Anthropic": PURPLE}
    for cost, quality, tokens, provider, workflow in points:
        x = x0 + ((cost - min_cost) / (max_cost - min_cost)) * chart_w if max_cost > min_cost else x0 + chart_w / 2
        y = y0 + chart_h - ((quality - (min_q - q_pad)) / ((max_q + q_pad) - (min_q - q_pad))) * chart_h
        radius = 15 + min(24, tokens / 160)
        d.ellipse((x - radius, y - radius, x + radius, y + radius), fill=colors.get(provider, RED), outline="white", width=3)
        label = f"{provider}\n{workflow.replace('_', ' ')}"
        d.multiline_text((x + radius + 6, y - 18), label, fill=INK, font=font(12), spacing=2)
    d.text((x0 + chart_w / 2 - 90, y0 + chart_h + 52), "Mean cost per run (USD)", fill=INK, font=font(18, True))
    d.text((20, y0 + chart_h / 2 - 20), "Mean quality", fill=INK, font=font(18, True))
    d.text((x0, y0 + chart_h + 18), fmt(min_cost, 4), fill=MUTED, font=font(14))
    d.text((x0 + chart_w - 70, y0 + chart_h + 18), fmt(max_cost, 4), fill=MUTED, font=font(14))
    d.text((x0 - 75, y0 + chart_h - 10), fmt(min_q - q_pad, 2), fill=MUTED, font=font(14))
    d.text((x0 - 75, y0 - 8), fmt(max_q + q_pad, 2), fill=MUTED, font=font(14))
    d.text((50, height - 70), "Bubble size approximates mean total token usage. Figure supports operational trade-off interpretation only.", fill=MUTED, font=font(16))
    im.save(FIG_DIR / "figure_03_cost_quality_tradeoff.png", quality=95)


def efficiency_ranking(rank_rows: List[Dict[str, Any]]) -> None:
    rows = [r for r in rank_rows if as_float(r.get("balanced_efficiency_index")) is not None]
    rows.sort(key=lambda r: as_float(r.get("balanced_efficiency_index")) or 0, reverse=True)
    rows = rows[:10]
    width, height = 1500, 900
    im = Image.new("RGB", (width, height), BG)
    d = ImageDraw.Draw(im)
    draw_title(d, "Figure 4. Top operational-efficiency configurations", "Balanced index combines quality ratio with mean cost, latency, and token multipliers.", width)
    x0, y0, chart_w, row_h = 650, 165, 650, 58
    max_v = max(as_float(r.get("balanced_efficiency_index")) or 0 for r in rows)
    for i, r in enumerate(rows):
        y = y0 + i * row_h
        label = f"{r.get(PROVIDER_COL)} / {r.get('Model')} / {str(r.get(WORKFLOW_COL)).replace('_', ' ')}"
        if len(label) > 62:
            label = label[:59] + "..."
        d.text((55, y + 14), f"{i+1}. {label}", fill=INK, font=font(16, True if i < 3 else False))
        v = as_float(r.get("balanced_efficiency_index")) or 0
        bar_w = int((v / max_v) * chart_w) if max_v else 0
        d.rounded_rectangle((x0, y + 8, x0 + bar_w, y + row_h - 10), radius=12, fill=BLUE if i < 3 else "#60A5FA")
        d.text((x0 + bar_w + 12, y + 16), fmt(v, 3), fill=INK, font=font(16, True))
    d.text((50, height - 82), "Index is descriptive, not a causal ranking. It normalizes observed mean quality against cost, duration, and token use.", fill=MUTED, font=font(16))
    im.save(FIG_DIR / "figure_04_operational_efficiency_ranking.png", quality=95)


def autosize(ws) -> None:
    for column_cells in ws.columns:
        max_length = 0
        col_letter = get_column_letter(column_cells[0].column)
        for cell in column_cells:
            if cell.value is not None:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max(max_length + 2, 10), 65)


def write_sheet(wb: Workbook, name: str, rows: List[Dict[str, Any]]) -> None:
    ws = wb.create_sheet(name)
    if not rows:
        ws.append(["No records"])
        return
    headers: List[str] = []
    for row in rows:
        for key in row:
            if key not in headers:
                headers.append(key)
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="D9EAF7", end_color="D9EAF7", fill_type="solid")
    for row in rows:
        ws.append([row.get(h) for h in headers])
    ws.freeze_panes = "A2"
    autosize(ws)


def compact_rows(rows: List[Dict[str, Any]], cols: List[str]) -> List[Dict[str, Any]]:
    return [{col: row.get(col) for col in cols} for row in rows]


def write_tables(data_rows: List[Dict[str, Any]], efficiency_rows: List[Dict[str, Any]], task_rows: List[Dict[str, Any]]) -> None:
    provider_workflow = summarize(data_rows, (PROVIDER_COL, WORKFLOW_COL))
    by_category = summarize(data_rows, (CATEGORY_COL,))
    by_difficulty = summarize(data_rows, (DIFFICULTY_COL,))
    provider_workflow.sort(key=lambda r: (PROVIDER_ORDER.index(r[PROVIDER_COL]), WORKFLOW_ORDER.index(r[WORKFLOW_COL])))
    by_category.sort(key=lambda r: CATEGORY_ORDER.index(r[CATEGORY_COL]) if r[CATEGORY_COL] in CATEGORY_ORDER else 99)
    by_difficulty.sort(key=lambda r: DIFFICULTY_ORDER.index(r[DIFFICULTY_COL]) if r[DIFFICULTY_COL] in DIFFICULTY_ORDER else 99)

    rank_rows = [r for r in efficiency_rows if as_float(r.get("balanced_efficiency_index")) is not None]
    rank_rows.sort(key=lambda r: as_float(r.get("balanced_efficiency_index")) or 0, reverse=True)

    wb = Workbook()
    active = wb.active
    if active is not None:
        wb.remove(active)
    write_sheet(wb, "Table_1_Provider_Workflow", compact_rows(provider_workflow, [PROVIDER_COL, WORKFLOW_COL, "N", f"mean_{QUALITY_COL}", f"sd_{QUALITY_COL}", f"mean_{CONFIDENCE_COL}", f"mean_{COST_COL}", f"mean_{DURATION_COL}", f"mean_{TOKENS_COL}"]))
    write_sheet(wb, "Table_2_Task_Category", compact_rows(by_category, [CATEGORY_COL, "N", f"mean_{QUALITY_COL}", f"sd_{QUALITY_COL}", f"mean_{CONFIDENCE_COL}", f"mean_{COST_COL}", f"mean_{DURATION_COL}", f"mean_{TOKENS_COL}"]))
    write_sheet(wb, "Table_3_Task_Difficulty", compact_rows(by_difficulty, [DIFFICULTY_COL, "N", f"mean_{QUALITY_COL}", f"sd_{QUALITY_COL}", f"mean_{CONFIDENCE_COL}", f"mean_{COST_COL}", f"mean_{DURATION_COL}", f"mean_{TOKENS_COL}"]))
    write_sheet(wb, "Table_4_Efficiency_Top10", compact_rows(rank_rows[:10], [PROVIDER_COL, "Model", WORKFLOW_COL, "N", "mean_quality_score", "mean_cost_usd", "mean_duration_sec", "mean_total_tokens", "balanced_efficiency_index", "cost_multiplier_vs_lowest_mean", "latency_multiplier_vs_lowest_mean", "token_multiplier_vs_lowest_mean"]))
    wb.save(TABLE_XLSX)

    lines = [
        "# Study 002 Publication Tables v2",
        "",
        "Public-safe generated tables for manuscript drafting. `quality_score` is interpreted as an operational workflow-generated quality proxy, not an independent human judgment.",
        "",
        "## Table 1. Provider × workflow descriptive summary",
        "",
    ]
    for row in compact_rows(provider_workflow, [PROVIDER_COL, WORKFLOW_COL, "N", f"mean_{QUALITY_COL}", f"sd_{QUALITY_COL}", f"mean_{COST_COL}", f"mean_{DURATION_COL}", f"mean_{TOKENS_COL}"]):
        lines.append(f"- {row[PROVIDER_COL]} / {row[WORKFLOW_COL]}: N={row['N']}; mean quality={fmt(row[f'mean_{QUALITY_COL}'])}; SD={fmt(row[f'sd_{QUALITY_COL}'])}; mean cost=${fmt(row[f'mean_{COST_COL}'],4)}; duration={fmt(row[f'mean_{DURATION_COL}'])}s; tokens={fmt(row[f'mean_{TOKENS_COL}'],1)}")
    lines.extend(["", "## Table 2. Task-category summary", ""])
    for row in by_category:
        lines.append(f"- {row[CATEGORY_COL]}: N={row['N']}; mean quality={fmt(row[f'mean_{QUALITY_COL}'])}; SD={fmt(row[f'sd_{QUALITY_COL}'])}; mean confidence={fmt(row[f'mean_{CONFIDENCE_COL}'])}")
    lines.extend(["", "## Table 3. Difficulty-annotation summary", "", "Difficulty was a secondary annotation layer and was not the primary task-bank balancing criterion.", ""])
    for row in by_difficulty:
        lines.append(f"- {row[DIFFICULTY_COL]}: N={row['N']}; mean quality={fmt(row[f'mean_{QUALITY_COL}'])}; SD={fmt(row[f'sd_{QUALITY_COL}'])}; mean confidence={fmt(row[f'mean_{CONFIDENCE_COL}'])}")
    lines.extend(["", "## Table 4. Top 10 operational-efficiency configurations", ""])
    for i, row in enumerate(rank_rows[:10], 1):
        lines.append(f"- {i}. {row.get(PROVIDER_COL)} / {row.get('Model')} / {row.get(WORKFLOW_COL)}: balanced efficiency index={fmt(row.get('balanced_efficiency_index'))}; mean quality={fmt(row.get('mean_quality_score'))}; mean cost=${fmt(row.get('mean_cost_usd'),4)}")
    TABLE_MD.write_text("\n".join(lines) + "\n", encoding="utf-8")


def main() -> None:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    FIG_DIR.mkdir(parents=True, exist_ok=True)
    rows = read_records(DATASET, DATA_SHEET)
    efficiency_rows = read_records(EFFICIENCY, "configuration_rankings")
    pw_rows = summarize(rows, (PROVIDER_COL, WORKFLOW_COL))
    pw_rows.sort(key=lambda r: (PROVIDER_ORDER.index(r[PROVIDER_COL]), WORKFLOW_ORDER.index(r[WORKFLOW_COL])))
    category_rows = summarize(rows, (CATEGORY_COL,))
    category_rows.sort(key=lambda r: CATEGORY_ORDER.index(r[CATEGORY_COL]) if r[CATEGORY_COL] in CATEGORY_ORDER else 99)

    heatmap_provider_workflow(pw_rows)
    grouped_bar_by_category(category_rows)
    cost_quality_tradeoff(pw_rows)
    efficiency_ranking(efficiency_rows)
    write_tables(rows, efficiency_rows, read_records(TASK_STRATIFIED, "Stratified_Summaries"))

    print(f"Generated figures in {FIG_DIR}")
    for path in sorted(FIG_DIR.glob("*.png")):
        print(f" - {path.name} ({path.stat().st_size} bytes)")
    print(f"Generated {TABLE_XLSX} ({TABLE_XLSX.stat().st_size} bytes)")
    print(f"Generated {TABLE_MD} ({TABLE_MD.stat().st_size} bytes)")


if __name__ == "__main__":
    main()
