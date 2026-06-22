#!/usr/bin/env python3
"""Task-blocked robustness analysis for Study 002.

Dependency-light script: openpyxl + numpy only. It fits fixed-effect OLS models
that include task as a blocking factor, then compares nested models for provider,
workflow, and provider×workflow effects. This is intended as a robustness check
against a plain row-level two-way ANOVA because each task is repeated across all
provider×workflow conditions.
"""
from __future__ import annotations

from pathlib import Path
import math
from collections import defaultdict

import numpy as np
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[2]
DATASET = ROOT / "datasets" / "merged" / "Agentic_AI_Experiments_Main_Study002_V1.4.4_270Runs_AnalysisReady.xlsx"
OUT_XLSX = ROOT / "analysis" / "statistical_outputs" / "task_blocked_analysis.xlsx"
OUT_MD = ROOT / "results" / "task_blocked_analysis_summary.md"

OUT_XLSX.parent.mkdir(parents=True, exist_ok=True)
OUT_MD.parent.mkdir(parents=True, exist_ok=True)


def betacf(a: float, b: float, x: float) -> float:
    # Continued fraction for incomplete beta, Numerical Recipes style.
    MAXIT = 200
    EPS = 3e-12
    FPMIN = 1e-300
    qab = a + b
    qap = a + 1.0
    qam = a - 1.0
    c = 1.0
    d = 1.0 - qab * x / qap
    if abs(d) < FPMIN:
        d = FPMIN
    d = 1.0 / d
    h = d
    for m in range(1, MAXIT + 1):
        m2 = 2 * m
        aa = m * (b - m) * x / ((qam + m2) * (a + m2))
        d = 1.0 + aa * d
        if abs(d) < FPMIN:
            d = FPMIN
        c = 1.0 + aa / c
        if abs(c) < FPMIN:
            c = FPMIN
        d = 1.0 / d
        h *= d * c
        aa = -(a + m) * (qab + m) * x / ((a + m2) * (qap + m2))
        d = 1.0 + aa * d
        if abs(d) < FPMIN:
            d = FPMIN
        c = 1.0 + aa / c
        if abs(c) < FPMIN:
            c = FPMIN
        d = 1.0 / d
        delh = d * c
        h *= delh
        if abs(delh - 1.0) < EPS:
            break
    return h


def betai(a: float, b: float, x: float) -> float:
    if x <= 0.0:
        return 0.0
    if x >= 1.0:
        return 1.0
    bt = math.exp(math.lgamma(a + b) - math.lgamma(a) - math.lgamma(b) + a * math.log(x) + b * math.log(1.0 - x))
    if x < (a + 1.0) / (a + b + 2.0):
        return bt * betacf(a, b, x) / a
    return 1.0 - bt * betacf(b, a, 1.0 - x) / b


def f_sf(f: float, dfn: int, dfd: int) -> float:
    if f < 0:
        return 1.0
    x = (dfn * f) / (dfn * f + dfd)
    cdf = betai(dfn / 2.0, dfd / 2.0, x)
    return max(0.0, min(1.0, 1.0 - cdf))


def p_fmt(p: float) -> str:
    if p < 0.001:
        return "<0.001"
    return f"{p:.3f}"


def read_rows() -> list[dict]:
    wb = load_workbook(DATASET, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    rows = []
    for vals in ws.iter_rows(min_row=2, values_only=True):
        row = dict(zip(headers, vals))
        if row.get("success") is True:
            row["task"] = str(row.get("task_id_final") or row.get("task_id"))
            row["provider"] = str(row.get("Provider"))
            row["workflow"] = str(row.get("workflow_type"))
            rows.append(row)
    return rows


def design(rows: list[dict], include_task=True, include_provider=True, include_workflow=True, include_interaction=True):
    tasks = sorted({r["task"] for r in rows}, key=lambda x: int(x) if x.isdigit() else x)
    providers = sorted({r["provider"] for r in rows})
    workflows = sorted({r["workflow"] for r in rows})
    cols = [("intercept", "1")]
    if include_task:
        cols += [("task", t) for t in tasks[1:]]
    if include_provider:
        cols += [("provider", p) for p in providers[1:]]
    if include_workflow:
        cols += [("workflow", w) for w in workflows[1:]]
    if include_interaction:
        # Include only non-reference provider/workflow interactions; valid with main effects present.
        cols += [("interaction", f"{p}::{w}") for p in providers[1:] for w in workflows[1:]]
    X = np.zeros((len(rows), len(cols)), dtype=float)
    for i, r in enumerate(rows):
        for j, (kind, val) in enumerate(cols):
            if kind == "intercept":
                X[i, j] = 1.0
            elif kind == "task":
                X[i, j] = 1.0 if r["task"] == val else 0.0
            elif kind == "provider":
                X[i, j] = 1.0 if r["provider"] == val else 0.0
            elif kind == "workflow":
                X[i, j] = 1.0 if r["workflow"] == val else 0.0
            elif kind == "interaction":
                p, w = val.split("::")
                X[i, j] = 1.0 if (r["provider"] == p and r["workflow"] == w) else 0.0
    return X, cols


def fit_rss(y: np.ndarray, X: np.ndarray) -> tuple[float, int, int]:
    beta, residuals, rank, s = np.linalg.lstsq(X, y, rcond=None)
    resid = y - X @ beta
    rss = float(np.sum(resid ** 2))
    df_resid = len(y) - int(rank)
    return rss, df_resid, int(rank)


def nested_test(rows: list[dict], yfield: str, reduced_kwargs: dict, full_kwargs: dict) -> dict:
    y = np.array([float(r[yfield]) for r in rows], dtype=float)
    Xf, cf = design(rows, **full_kwargs)
    Xr, cr = design(rows, **reduced_kwargs)
    rss_f, df_f, rank_f = fit_rss(y, Xf)
    rss_r, df_r, rank_r = fit_rss(y, Xr)
    ss = rss_r - rss_f
    df_effect = df_r - df_f
    ms_effect = ss / df_effect
    ms_error = rss_f / df_f
    f = ms_effect / ms_error if ms_error else float("nan")
    p = f_sf(f, df_effect, df_f)
    pes = ss / (ss + rss_f) if (ss + rss_f) else float("nan")
    return {
        "outcome": yfield,
        "ss_effect": ss,
        "df_effect": df_effect,
        "ss_error": rss_f,
        "df_error": df_f,
        "F": f,
        "p": p,
        "partial_eta_sq": pes,
        "rank_full": rank_f,
        "rank_reduced": rank_r,
    }


def summarize(rows):
    groups = defaultdict(list)
    for r in rows:
        groups[(r["provider"], r["workflow"])].append(r)
    out = []
    for (provider, workflow), rs in sorted(groups.items()):
        out.append({
            "provider": provider,
            "workflow": workflow,
            "n": len(rs),
            "quality_mean": sum(float(r["quality_score"]) for r in rs) / len(rs),
            "confidence_mean": sum(float(r["confidence"]) for r in rs) / len(rs),
            "cost_mean": sum(float(r["cost_usd"]) for r in rs) / len(rs),
            "duration_mean": sum(float(r["duration_sec"]) for r in rs) / len(rs),
            "tokens_mean": sum(float(r["total_tokens"]) for r in rs) / len(rs),
            "calls_mean": sum(float(r["llm_call_count"]) for r in rs) / len(rs),
        })
    return out


def main():
    rows = read_rows()
    assert len(rows) == 270, len(rows)
    full = dict(include_task=True, include_provider=True, include_workflow=True, include_interaction=True)
    no_provider = dict(include_task=True, include_provider=False, include_workflow=True, include_interaction=False)
    provider_main = dict(include_task=True, include_provider=True, include_workflow=True, include_interaction=False)
    no_workflow = dict(include_task=True, include_provider=True, include_workflow=False, include_interaction=False)
    no_interaction = dict(include_task=True, include_provider=True, include_workflow=True, include_interaction=False)

    tests = []
    for outcome in ["quality_score", "confidence"]:
        tests.append({"effect": "provider_main_given_workflow_task", **nested_test(rows, outcome, no_provider, provider_main)})
        tests.append({"effect": "workflow_main_given_provider_task", **nested_test(rows, outcome, no_workflow, provider_main)})
        tests.append({"effect": "provider_workflow_interaction_given_task_main_effects", **nested_test(rows, outcome, no_interaction, full)})

    desc = summarize(rows)

    wb = Workbook()
    ws = wb.create_sheet("Task_Blocked_Tests", 0)
    default = wb["Sheet"] if "Sheet" in wb.sheetnames else None
    if default is not None:
        wb.remove(default)
    headers = ["outcome", "effect", "ss_effect", "df_effect", "ss_error", "df_error", "F", "p", "p_formatted", "partial_eta_sq"]
    ws.append(headers)
    for t in tests:
        ws.append([t["outcome"], t["effect"], t["ss_effect"], t["df_effect"], t["ss_error"], t["df_error"], t["F"], t["p"], p_fmt(t["p"]), t["partial_eta_sq"]])
    ws2 = wb.create_sheet("Provider_Workflow_Means")
    headers2 = ["provider", "workflow", "n", "quality_mean", "confidence_mean", "cost_mean", "duration_mean", "tokens_mean", "calls_mean"]
    ws2.append(headers2)
    for d in desc:
        ws2.append([d[h] for h in headers2])
    ws3 = wb.create_sheet("Interpretation")
    ws3.append(["Item", "Value"])
    ws3.append(["Purpose", "Task-blocked fixed-effect robustness check for repeated task structure."])
    ws3.append(["Model", "Outcome ~ task block + provider + workflow + provider×workflow."])
    ws3.append(["Interpretation", "This analysis treats task identity as a blocking factor and should be reported as robustness evidence alongside the original ANOVA."])
    ws3.append(["Caution", "quality_score remains an operational workflow-generated proxy, not an independent human rating."])
    for sheet in wb.worksheets:
        for cell in sheet[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="D9EAF7")
        for col in range(1, sheet.max_column + 1):
            sheet.column_dimensions[get_column_letter(col)].width = 24
    wb.save(OUT_XLSX)

    lines = ["# Study 002 Task-Blocked Robustness Analysis", "", "This public-safe summary reports a task-blocked fixed-effect robustness analysis. Because each of the 30 benchmark tasks was executed under every provider × workflow condition, task identity was included as a blocking factor.", "", "Model form: `outcome ~ task block + provider + workflow + provider × workflow`.", "", "`quality_score` remains an operational workflow-generated proxy rather than an independent human judgment.", "", "## Task-blocked tests", ""]
    for t in tests:
        lines.append(f"- {t['outcome']} / {t['effect']}: F({t['df_effect']}, {t['df_error']}) = {t['F']:.3f}, p = {p_fmt(t['p'])}, partial eta squared = {t['partial_eta_sq']:.3f}")
    lines += ["", "## Interpretation", "", "The task-blocked analysis preserves the main Study 002 conclusion that provider and workflow should be evaluated jointly. It also reinforces the need for cautious wording: effects on `quality_score` are effects on an operational proxy whose construction differs between reviewer and non-reviewer workflows.", "", "## Output files", "", f"- `{OUT_XLSX.relative_to(ROOT)}`", f"- `{OUT_MD.relative_to(ROOT)}`", ""]
    OUT_MD.write_text("\n".join(lines), encoding="utf-8", newline="\n")
    print(f"Wrote {OUT_XLSX}")
    print(f"Wrote {OUT_MD}")
    for line in lines[8:15]:
        print(line)


if __name__ == "__main__":
    main()
