#!/usr/bin/env python3
"""Study 002 operational efficiency analysis.

This dependency-light script summarizes Study 002 efficiency metrics for the
public reproducibility package. The analysis treats `quality_score` as an
operational quality proxy and reports observed efficiency associations, not
causal performance claims.

Outputs:
  analysis/statistical_outputs/operational_efficiency.xlsx

Analyses included:
  - quality proxy per dollar, per 1k tokens, and per second
  - confidence per dollar, per 1k tokens, and per second
  - cost, latency, and token multipliers against the most efficient observed
    configuration baseline
  - provider, workflow, and provider×workflow summaries
  - configuration rankings across quality, cost, latency, token use, and
    composite operational efficiency
"""

from __future__ import annotations

from collections import defaultdict
from math import sqrt
from pathlib import Path
from statistics import mean, median
from typing import Any, Dict, Iterable, List, Sequence, Tuple

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
STUDY_ROOT = ROOT.parent
DATASET = STUDY_ROOT / "datasets" / "merged" / "Agentic_AI_Experiments_Main_Study002_V1.4.4_270Runs_AnalysisReady.xlsx"
OUTPUT = STUDY_ROOT / "analysis" / "statistical_outputs" / "operational_efficiency.xlsx"

DATA_SHEET = "Agentic_AI_Experiments_V1.0"
PROVIDER_COL = "Provider"
MODEL_COL = "Model"
WORKFLOW_COL = "workflow_type"
QUALITY_COL = "quality_score"
CONFIDENCE_COL = "confidence"
COST_COL = "cost_usd"
DURATION_COL = "duration_sec"
TOKENS_COL = "total_tokens"
CALLS_COL = "llm_call_count"
SUCCESS_COL = "success"

GROUPINGS: Tuple[Tuple[str, Tuple[str, ...]], ...] = (
    ("provider", (PROVIDER_COL,)),
    ("workflow", (WORKFLOW_COL,)),
    ("provider_workflow", (PROVIDER_COL, WORKFLOW_COL)),
    ("provider_model_workflow", (PROVIDER_COL, MODEL_COL, WORKFLOW_COL)),
)


Number = int | float


def as_float(value: Any) -> float | None:
    if value is None:
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def safe_divide(numerator: float | None, denominator: float | None) -> float | None:
    if numerator is None or denominator is None or denominator <= 0:
        return None
    return numerator / denominator


def safe_mean(values: Iterable[float]) -> float | None:
    vals = list(values)
    return mean(vals) if vals else None


def percentile(values: Sequence[float], p: float) -> float | None:
    vals = sorted(values)
    if not vals:
        return None
    if len(vals) == 1:
        return vals[0]
    rank = (len(vals) - 1) * p
    lower = int(rank)
    upper = min(lower + 1, len(vals) - 1)
    weight = rank - lower
    return vals[lower] * (1 - weight) + vals[upper] * weight


def stddev(values: Sequence[float]) -> float | None:
    vals = list(values)
    if len(vals) < 2:
        return None
    m = mean(vals)
    return sqrt(sum((x - m) ** 2 for x in vals) / (len(vals) - 1))


def read_rows() -> List[Dict[str, Any]]:
    if not DATASET.exists():
        raise FileNotFoundError(f"Dataset not found: {DATASET}")
    wb = load_workbook(DATASET, read_only=True, data_only=True)
    if DATA_SHEET not in wb.sheetnames:
        raise ValueError(f"Sheet {DATA_SHEET!r} not found. Available sheets: {wb.sheetnames}")
    ws = wb[DATA_SHEET]
    headers = [str(h).strip() if h is not None else "" for h in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
    required = [PROVIDER_COL, MODEL_COL, WORKFLOW_COL, QUALITY_COL, CONFIDENCE_COL, COST_COL, DURATION_COL, TOKENS_COL]
    missing = [col for col in required if col not in headers]
    if missing:
        raise ValueError(f"Missing required columns: {missing}")

    rows: List[Dict[str, Any]] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        record = dict(zip(headers, row))
        if record.get(PROVIDER_COL) and record.get(WORKFLOW_COL):
            rows.append(enrich_row(record))
    return rows


def enrich_row(record: Dict[str, Any]) -> Dict[str, Any]:
    row = dict(record)
    quality = as_float(row.get(QUALITY_COL))
    confidence = as_float(row.get(CONFIDENCE_COL))
    cost = as_float(row.get(COST_COL))
    duration = as_float(row.get(DURATION_COL))
    tokens = as_float(row.get(TOKENS_COL))
    token_units = (tokens / 1000.0) if tokens is not None else None

    row["quality_per_usd"] = safe_divide(quality, cost)
    row["quality_per_1k_tokens"] = safe_divide(quality, token_units)
    row["quality_per_second"] = safe_divide(quality, duration)
    row["confidence_per_usd"] = safe_divide(confidence, cost)
    row["confidence_per_1k_tokens"] = safe_divide(confidence, token_units)
    row["confidence_per_second"] = safe_divide(confidence, duration)
    return row


def group_rows(rows: List[Dict[str, Any]], columns: Tuple[str, ...]) -> Dict[Tuple[Any, ...], List[Dict[str, Any]]]:
    grouped: Dict[Tuple[Any, ...], List[Dict[str, Any]]] = defaultdict(list)
    for row in rows:
        grouped[tuple(row.get(col) for col in columns)].append(row)
    return grouped


def values_for(rows: List[Dict[str, Any]], metric: str) -> List[float]:
    return [value for row in rows if (value := as_float(row.get(metric))) is not None]


def summarize_group(rows: List[Dict[str, Any]], grouping_name: str, columns: Tuple[str, ...]) -> List[Dict[str, Any]]:
    records: List[Dict[str, Any]] = []
    grouped = group_rows(rows, columns)
    for key in sorted(grouped, key=lambda parts: tuple(str(part) for part in parts)):
        group = grouped[key]
        record: Dict[str, Any] = {"grouping": grouping_name}
        record.update({col: value for col, value in zip(columns, key)})
        record["N"] = len(group)
        if SUCCESS_COL in group[0]:
            success_values = [str(r.get(SUCCESS_COL)).strip().lower() for r in group if r.get(SUCCESS_COL) is not None]
            record["success_rate"] = (sum(v in {"true", "1", "yes", "success"} for v in success_values) / len(success_values)) if success_values else None

        metrics = [
            QUALITY_COL,
            CONFIDENCE_COL,
            COST_COL,
            DURATION_COL,
            TOKENS_COL,
            CALLS_COL,
            "quality_per_usd",
            "quality_per_1k_tokens",
            "quality_per_second",
            "confidence_per_usd",
            "confidence_per_1k_tokens",
            "confidence_per_second",
        ]
        for metric in metrics:
            vals = values_for(group, metric)
            record[f"mean_{metric}"] = mean(vals) if vals else None
            record[f"median_{metric}"] = median(vals) if vals else None
            record[f"sd_{metric}"] = stddev(vals)
            record[f"q1_{metric}"] = percentile(vals, 0.25)
            record[f"q3_{metric}"] = percentile(vals, 0.75)
            record[f"min_{metric}"] = min(vals) if vals else None
            record[f"max_{metric}"] = max(vals) if vals else None
        records.append(record)
    return records


def add_multipliers(records: List[Dict[str, Any]]) -> None:
    for grouping in sorted({str(r.get("grouping")) for r in records}):
        scoped = [r for r in records if r.get("grouping") == grouping]
        min_cost = min((r["mean_cost_usd"] for r in scoped if r.get("mean_cost_usd") is not None and r["mean_cost_usd"] > 0), default=None)
        min_latency = min((r["mean_duration_sec"] for r in scoped if r.get("mean_duration_sec") is not None and r["mean_duration_sec"] > 0), default=None)
        min_tokens = min((r["mean_total_tokens"] for r in scoped if r.get("mean_total_tokens") is not None and r["mean_total_tokens"] > 0), default=None)
        max_quality = max((r["mean_quality_score"] for r in scoped if r.get("mean_quality_score") is not None), default=None)
        for record in scoped:
            record["cost_multiplier_vs_lowest_mean"] = safe_divide(record.get("mean_cost_usd"), min_cost)
            record["latency_multiplier_vs_lowest_mean"] = safe_divide(record.get("mean_duration_sec"), min_latency)
            record["token_multiplier_vs_lowest_mean"] = safe_divide(record.get("mean_total_tokens"), min_tokens)
            record["quality_ratio_vs_highest_mean"] = safe_divide(record.get("mean_quality_score"), max_quality)
            c = record.get("cost_multiplier_vs_lowest_mean")
            l = record.get("latency_multiplier_vs_lowest_mean")
            t = record.get("token_multiplier_vs_lowest_mean")
            q = record.get("quality_ratio_vs_highest_mean")
            efficiency_penalty = mean([x for x in (c, l, t) if x is not None]) if any(x is not None for x in (c, l, t)) else None
            record["balanced_efficiency_index"] = safe_divide(q, efficiency_penalty)


def rank_desc(records: List[Dict[str, Any]], metric: str) -> Dict[int, int | None]:
    valid = [(idx, as_float(record.get(metric))) for idx, record in enumerate(records)]
    valid = [(idx, value) for idx, value in valid if value is not None]
    valid.sort(key=lambda item: item[1], reverse=True)
    ranks: Dict[int, int | None] = {idx: None for idx in range(len(records))}
    for rank, (idx, _) in enumerate(valid, start=1):
        ranks[idx] = rank
    return ranks


def rank_asc(records: List[Dict[str, Any]], metric: str) -> Dict[int, int | None]:
    valid = [(idx, as_float(record.get(metric))) for idx, record in enumerate(records)]
    valid = [(idx, value) for idx, value in valid if value is not None]
    valid.sort(key=lambda item: item[1])
    ranks: Dict[int, int | None] = {idx: None for idx in range(len(records))}
    for rank, (idx, _) in enumerate(valid, start=1):
        ranks[idx] = rank
    return ranks


def ranking_records(summary_records: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    configs = [r for r in summary_records if r.get("grouping") == "provider_model_workflow"]
    rank_maps = {
        "quality_rank_desc": rank_desc(configs, "mean_quality_score"),
        "quality_per_usd_rank_desc": rank_desc(configs, "mean_quality_per_usd"),
        "quality_per_1k_tokens_rank_desc": rank_desc(configs, "mean_quality_per_1k_tokens"),
        "quality_per_second_rank_desc": rank_desc(configs, "mean_quality_per_second"),
        "lowest_cost_rank_asc": rank_asc(configs, "mean_cost_usd"),
        "lowest_latency_rank_asc": rank_asc(configs, "mean_duration_sec"),
        "lowest_token_rank_asc": rank_asc(configs, "mean_total_tokens"),
        "balanced_efficiency_rank_desc": rank_desc(configs, "balanced_efficiency_index"),
    }
    records: List[Dict[str, Any]] = []
    for idx, config in enumerate(configs):
        rank_values = [rank for name in rank_maps if (rank := rank_maps[name][idx]) is not None]
        record = {
            PROVIDER_COL: config.get(PROVIDER_COL),
            MODEL_COL: config.get(MODEL_COL),
            WORKFLOW_COL: config.get(WORKFLOW_COL),
            "N": config.get("N"),
            "mean_quality_score": config.get("mean_quality_score"),
            "mean_confidence": config.get("mean_confidence"),
            "mean_cost_usd": config.get("mean_cost_usd"),
            "mean_duration_sec": config.get("mean_duration_sec"),
            "mean_total_tokens": config.get("mean_total_tokens"),
            "mean_quality_per_usd": config.get("mean_quality_per_usd"),
            "mean_quality_per_1k_tokens": config.get("mean_quality_per_1k_tokens"),
            "mean_quality_per_second": config.get("mean_quality_per_second"),
            "cost_multiplier_vs_lowest_mean": config.get("cost_multiplier_vs_lowest_mean"),
            "latency_multiplier_vs_lowest_mean": config.get("latency_multiplier_vs_lowest_mean"),
            "token_multiplier_vs_lowest_mean": config.get("token_multiplier_vs_lowest_mean"),
            "balanced_efficiency_index": config.get("balanced_efficiency_index"),
            "mean_rank_across_reported_efficiency_criteria": mean(rank_values) if rank_values else None,
        }
        for rank_name, ranks in rank_maps.items():
            record[rank_name] = ranks[idx]
        records.append(record)
    records.sort(key=lambda r: (r["mean_rank_across_reported_efficiency_criteria"] is None, r["mean_rank_across_reported_efficiency_criteria"] or 9999))
    return records


def overall_summary(rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    metrics = [QUALITY_COL, CONFIDENCE_COL, COST_COL, DURATION_COL, TOKENS_COL, "quality_per_usd", "quality_per_1k_tokens", "quality_per_second"]
    records: List[Dict[str, Any]] = []
    for metric in metrics:
        vals = values_for(rows, metric)
        records.append(
            {
                "metric": metric,
                "N": len(vals),
                "mean": mean(vals) if vals else None,
                "median": median(vals) if vals else None,
                "sd": stddev(vals),
                "q1": percentile(vals, 0.25),
                "q3": percentile(vals, 0.75),
                "min": min(vals) if vals else None,
                "max": max(vals) if vals else None,
            }
        )
    return records


def metadata_records(rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    return [
        {"field": "dataset", "value": str(DATASET.relative_to(STUDY_ROOT))},
        {"field": "output", "value": str(OUTPUT.relative_to(STUDY_ROOT))},
        {"field": "row_count", "value": len(rows)},
        {"field": "provider_count", "value": len({r.get(PROVIDER_COL) for r in rows})},
        {"field": "workflow_count", "value": len({r.get(WORKFLOW_COL) for r in rows})},
        {"field": "analysis_framing", "value": "Operational efficiency summary using workflow-generated quality proxy; exploratory and descriptive."},
        {"field": "multiplier_baseline", "value": "Lowest observed mean cost, latency, or token count within the same grouping sheet."},
        {"field": "ranking_scope", "value": "Provider-model-workflow configurations only; ranks are descriptive, not inferential."},
    ]


def autosize(ws) -> None:
    for col_idx, column_cells in enumerate(ws.columns, start=1):
        max_len = 0
        for cell in column_cells:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 12), 70)


def write_sheet(wb: Workbook, name: str, records: List[Dict[str, Any]]) -> None:
    ws = wb.create_sheet(name)
    if not records:
        ws.append(["No records"])
        return
    headers = list(records[0].keys())
    ws.append(headers)
    for record in records:
        ws.append([record.get(header) for header in headers])

    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
    ws.freeze_panes = "A2"
    autosize(ws)


def build_workbook(rows: List[Dict[str, Any]]) -> None:
    wb = Workbook()
    active_sheet = wb.active
    if active_sheet is not None:
        wb.remove(active_sheet)

    summary_records: List[Dict[str, Any]] = []
    for grouping_name, columns in GROUPINGS:
        grouped_records = summarize_group(rows, grouping_name, columns)
        add_multipliers(grouped_records)
        summary_records.extend(grouped_records)
        write_sheet(wb, f"summary_{grouping_name}"[:31], grouped_records)

    write_sheet(wb, "configuration_rankings", ranking_records(summary_records))
    write_sheet(wb, "overall_summary", overall_summary(rows))
    write_sheet(wb, "metadata", metadata_records(rows))

    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT)


def main() -> None:
    rows = read_rows()
    build_workbook(rows)
    print(f"Operational efficiency analysis written to {OUTPUT}")
    print(f"Rows analyzed: {len(rows)}")


if __name__ == "__main__":
    main()
