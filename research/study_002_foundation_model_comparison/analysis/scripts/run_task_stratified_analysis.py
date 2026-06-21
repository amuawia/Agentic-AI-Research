#!/usr/bin/env python3
"""Study 002 task-stratified analysis.

This dependency-light script summarizes Study 002 outcomes by task category and
secondary difficulty annotations. It treats `quality_score` as an operational
workflow-generated quality proxy rather than an independent human judgment.

Outputs:
  analysis/statistical_outputs/task_stratified_analysis.xlsx

Analyses included:
  - task-bank balance checks by category and difficulty
  - descriptive metrics by category, difficulty, category×difficulty
  - category/workflow and category/provider interaction summaries
  - difficulty/workflow and difficulty/provider summaries
  - per-task descriptive summaries for reproducibility
"""

from __future__ import annotations

from collections import Counter, defaultdict
from math import sqrt
from pathlib import Path
from statistics import mean, median
from typing import Any, Dict, Iterable, List, Sequence, Tuple

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
STUDY_ROOT = ROOT.parent
DATASET = STUDY_ROOT / "datasets" / "merged" / "Agentic_AI_Experiments_Main_Study002_V1.4.4_270Runs_AnalysisReady.xlsx"
OUTPUT = STUDY_ROOT / "analysis" / "statistical_outputs" / "task_stratified_analysis.xlsx"

DATA_SHEET = "Agentic_AI_Experiments_V1.0"
TASK_BANK_SHEET = "task_bank_v1"

TASK_ID_COL = "task_id_final"
CATEGORY_COL = "task_category_final"
DIFFICULTY_COL = "task_difficulty_final"
PROVIDER_COL = "Provider"
MODEL_COL = "Model"
WORKFLOW_COL = "workflow_type"
QUALITY_COL = "quality_score"
CONFIDENCE_COL = "confidence"
COST_COL = "cost_usd"
DURATION_COL = "duration_sec"
TOKENS_COL = "total_tokens"
CALLS_COL = "llm_call_count"
SUCCESS_COL = "success"

METRICS = (
    QUALITY_COL,
    CONFIDENCE_COL,
    COST_COL,
    DURATION_COL,
    TOKENS_COL,
    CALLS_COL,
)

GROUPINGS: Tuple[Tuple[str, Tuple[str, ...]], ...] = (
    ("category", (CATEGORY_COL,)),
    ("difficulty", (DIFFICULTY_COL,)),
    ("category_difficulty", (CATEGORY_COL, DIFFICULTY_COL)),
    ("category_workflow", (CATEGORY_COL, WORKFLOW_COL)),
    ("category_provider", (CATEGORY_COL, PROVIDER_COL)),
    ("difficulty_workflow", (DIFFICULTY_COL, WORKFLOW_COL)),
    ("difficulty_provider", (DIFFICULTY_COL, PROVIDER_COL)),
    ("category_provider_workflow", (CATEGORY_COL, PROVIDER_COL, WORKFLOW_COL)),
    ("difficulty_provider_workflow", (DIFFICULTY_COL, PROVIDER_COL, WORKFLOW_COL)),
)


def as_float(value: Any) -> float | None:
    if value is None:
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def safe_mean(values: Iterable[float]) -> float | None:
    vals = list(values)
    return mean(vals) if vals else None


def percentile(values: Sequence[float], p: float) -> float | None:
    vals = sorted(values)
    if not vals:
        return None
    if len(vals) == 1:
        return vals[0]
    rank = (len(vals) - 1) * p
    lower = int(rank)
    upper = min(lower + 1, len(vals) - 1)
    weight = rank - lower
    return vals[lower] * (1 - weight) + vals[upper] * weight


def stddev(values: Sequence[float]) -> float | None:
    vals = list(values)
    if len(vals) < 2:
        return None
    m = mean(vals)
    return sqrt(sum((x - m) ** 2 for x in vals) / (len(vals) - 1))


def normalize_success(value: Any) -> bool | None:
    if value is None:
        return None
    text = str(value).strip().lower()
    if text in {"true", "1", "yes", "success"}:
        return True
    if text in {"false", "0", "no", "failure", "failed"}:
        return False
    return None


def read_sheet_records(sheet_name: str) -> List[Dict[str, Any]]:
    if not DATASET.exists():
        raise FileNotFoundError(f"Dataset not found: {DATASET}")
    wb = load_workbook(DATASET, read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet {sheet_name!r} not found. Available sheets: {wb.sheetnames}")
    ws = wb[sheet_name]
    headers = [str(h).strip() if h is not None else "" for h in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
    rows: List[Dict[str, Any]] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        record = dict(zip(headers, row))
        if any(value is not None for value in record.values()):
            rows.append(record)
    return rows


def read_analysis_rows() -> List[Dict[str, Any]]:
    rows = read_sheet_records(DATA_SHEET)
    required = [TASK_ID_COL, CATEGORY_COL, DIFFICULTY_COL, PROVIDER_COL, WORKFLOW_COL, QUALITY_COL, CONFIDENCE_COL]
    missing = [col for col in required if col not in rows[0]] if rows else required
    if missing:
        raise ValueError(f"Missing required analysis columns: {missing}")
    return [row for row in rows if row.get(TASK_ID_COL) is not None and row.get(CATEGORY_COL) and row.get(DIFFICULTY_COL)]


def read_task_bank_rows() -> List[Dict[str, Any]]:
    rows = read_sheet_records(TASK_BANK_SHEET)
    required = ["task_id", "category", "difficulty", "task"]
    missing = [col for col in required if col not in rows[0]] if rows else required
    if missing:
        raise ValueError(f"Missing required task-bank columns: {missing}")
    return rows


def values_for(rows: List[Dict[str, Any]], metric: str) -> List[float]:
    return [value for row in rows if (value := as_float(row.get(metric))) is not None]


def group_rows(rows: List[Dict[str, Any]], columns: Tuple[str, ...]) -> Dict[Tuple[Any, ...], List[Dict[str, Any]]]:
    grouped: Dict[Tuple[Any, ...], List[Dict[str, Any]]] = defaultdict(list)
    for row in rows:
        grouped[tuple(row.get(col) for col in columns)].append(row)
    return grouped


def summarize_rows(rows: List[Dict[str, Any]], grouping_name: str, columns: Tuple[str, ...]) -> List[Dict[str, Any]]:
    records: List[Dict[str, Any]] = []
    grouped = group_rows(rows, columns)
    for key in sorted(grouped, key=lambda parts: tuple(str(part) for part in parts)):
        group = grouped[key]
        record: Dict[str, Any] = {"grouping": grouping_name}
        record.update({col: value for col, value in zip(columns, key)})
        record["N_runs"] = len(group)
        record["N_tasks"] = len({str(row.get(TASK_ID_COL)) for row in group if row.get(TASK_ID_COL) is not None})
        success_values = [normalize_success(row.get(SUCCESS_COL)) for row in group]
        success_values = [value for value in success_values if value is not None]
        record["success_rate"] = (sum(success_values) / len(success_values)) if success_values else None
        for metric in METRICS:
            vals = values_for(group, metric)
            record[f"mean_{metric}"] = mean(vals) if vals else None
            record[f"median_{metric}"] = median(vals) if vals else None
            record[f"sd_{metric}"] = stddev(vals)
            record[f"q1_{metric}"] = percentile(vals, 0.25)
            record[f"q3_{metric}"] = percentile(vals, 0.75)
            record[f"min_{metric}"] = min(vals) if vals else None
            record[f"max_{metric}"] = max(vals) if vals else None
        records.append(record)
    return records


def task_balance_records(task_bank_rows: List[Dict[str, Any]], analysis_rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    bank_category = Counter(str(r.get("category")) for r in task_bank_rows)
    bank_difficulty = Counter(str(r.get("difficulty")) for r in task_bank_rows)
    bank_cat_diff = Counter((str(r.get("category")), str(r.get("difficulty"))) for r in task_bank_rows)

    analysis_task_ids = {str(row.get(TASK_ID_COL)) for row in analysis_rows if row.get(TASK_ID_COL) is not None}
    records: List[Dict[str, Any]] = [
        {
            "section": "design_note",
            "dimension": "task_bank_design",
            "level": "category-balanced",
            "task_count": len(task_bank_rows),
            "run_count": len(analysis_rows),
            "distinct_tasks_in_analysis": len(analysis_task_ids),
            "interpretation": "The accepted task bank uses 30 tasks balanced by category: Knowledge, Reasoning, and Coding. Difficulty labels are secondary annotations and were not the primary balancing criterion.",
        }
    ]
    for category, count in sorted(bank_category.items()):
        records.append({"section": "task_bank", "dimension": "category", "level": category, "task_count": count})
    for difficulty, count in sorted(bank_difficulty.items()):
        records.append({"section": "task_bank", "dimension": "difficulty_secondary", "level": difficulty, "task_count": count})
    for (category, difficulty), count in sorted(bank_cat_diff.items()):
        records.append({"section": "task_bank", "dimension": "category_by_secondary_difficulty", "level": f"{category} × {difficulty}", "task_count": count})
    return records


def per_task_records(rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    records: List[Dict[str, Any]] = []
    grouped = group_rows(rows, (TASK_ID_COL, CATEGORY_COL, DIFFICULTY_COL))
    for key in sorted(grouped, key=lambda parts: (int(parts[0]) if str(parts[0]).isdigit() else str(parts[0]), str(parts[1]), str(parts[2]))):
        group = grouped[key]
        record: Dict[str, Any] = {
            TASK_ID_COL: key[0],
            CATEGORY_COL: key[1],
            DIFFICULTY_COL: key[2],
            "N_runs": len(group),
            "N_providers": len({row.get(PROVIDER_COL) for row in group}),
            "N_workflows": len({row.get(WORKFLOW_COL) for row in group}),
        }
        for metric in METRICS:
            vals = values_for(group, metric)
            record[f"mean_{metric}"] = mean(vals) if vals else None
            record[f"median_{metric}"] = median(vals) if vals else None
            record[f"sd_{metric}"] = stddev(vals)
            record[f"min_{metric}"] = min(vals) if vals else None
            record[f"max_{metric}"] = max(vals) if vals else None
        records.append(record)
    return records


def autosize(ws) -> None:
    for column_cells in ws.columns:
        max_length = 0
        col_letter = get_column_letter(column_cells[0].column)
        for cell in column_cells:
            value = cell.value
            if value is not None:
                max_length = max(max_length, len(str(value)))
        ws.column_dimensions[col_letter].width = min(max(max_length + 2, 10), 70)


def write_records(wb: Workbook, sheet_name: str, records: List[Dict[str, Any]]) -> None:
    ws = wb.create_sheet(sheet_name)
    if not records:
        ws.append(["No records"])
        return
    headers: List[str] = []
    for record in records:
        for key in record:
            if key not in headers:
                headers.append(key)
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="D9EAF7", end_color="D9EAF7", fill_type="solid")
    for record in records:
        ws.append([record.get(header) for header in headers])
    ws.freeze_panes = "A2"
    autosize(ws)


def overview_records(rows: List[Dict[str, Any]], task_bank_rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    return [
        {"item": "analysis_runs", "value": len(rows)},
        {"item": "task_bank_tasks", "value": len(task_bank_rows)},
        {"item": "distinct_tasks_in_analysis", "value": len({str(row.get(TASK_ID_COL)) for row in rows})},
        {"item": "providers", "value": ", ".join(sorted({str(row.get(PROVIDER_COL)) for row in rows}))},
        {"item": "workflows", "value": ", ".join(sorted({str(row.get(WORKFLOW_COL)) for row in rows}))},
        {"item": "categories", "value": ", ".join(sorted({str(row.get(CATEGORY_COL)) for row in rows}))},
        {"item": "difficulty_annotation", "value": "Difficulty is a secondary annotation layer; the task bank was primarily balanced by category."},
        {"item": "quality_interpretation", "value": "quality_score is interpreted as an operational workflow-generated quality proxy, not an independent human judgment."},
    ]


def main() -> None:
    rows = read_analysis_rows()
    task_bank_rows = read_task_bank_rows()
    all_summary_records: List[Dict[str, Any]] = []
    for grouping_name, columns in GROUPINGS:
        all_summary_records.extend(summarize_rows(rows, grouping_name, columns))

    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    active = wb.active
    if active is not None:
        wb.remove(active)
    write_records(wb, "Overview", overview_records(rows, task_bank_rows))
    write_records(wb, "Task_Bank_Balance", task_balance_records(task_bank_rows, rows))
    write_records(wb, "Stratified_Summaries", all_summary_records)
    write_records(wb, "Per_Task", per_task_records(rows))

    # Convenience sheets for the main manuscript tables.
    write_records(wb, "By_Category", summarize_rows(rows, "category", (CATEGORY_COL,)))
    write_records(wb, "By_Difficulty", summarize_rows(rows, "difficulty", (DIFFICULTY_COL,)))
    write_records(wb, "Category_Workflow", summarize_rows(rows, "category_workflow", (CATEGORY_COL, WORKFLOW_COL)))
    write_records(wb, "Difficulty_Workflow", summarize_rows(rows, "difficulty_workflow", (DIFFICULTY_COL, WORKFLOW_COL)))
    wb.save(OUTPUT)
    print(f"Wrote {OUTPUT}")
    print(f"Rows analyzed: {len(rows)}")
    print(f"Tasks in task bank: {len(task_bank_rows)}")
    print(f"Distinct tasks in analysis: {len({str(row.get(TASK_ID_COL)) for row in rows})}")


if __name__ == "__main__":
    main()
