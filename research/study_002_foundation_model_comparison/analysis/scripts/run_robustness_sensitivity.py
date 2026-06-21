#!/usr/bin/env python3
"""Study 002 robustness and sensitivity analysis.

This dependency-light script generates robustness summaries for Study 002 using
only Python's standard library plus openpyxl. It is intended for public
reproducibility in minimal review environments.

Outputs:
  analysis/statistical_outputs/robustness_sensitivity.xlsx

Analyses included:
  - mean, median, Q1, Q3, IQR, SD, min, max by workflow/provider/provider×workflow
  - deterministic bootstrap confidence intervals for mean quality/confidence
  - IQR outlier sensitivity summaries
  - reviewer vs non-reviewer workflow comparison
  - simple mean-difference contrasts within provider and within workflow
"""

from __future__ import annotations

from collections import defaultdict
from math import sqrt
from pathlib import Path
from random import Random
from statistics import mean, median
from typing import Any, Dict, Iterable, List, Sequence, Tuple

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
STUDY_ROOT = ROOT.parent
DATASET = STUDY_ROOT / "datasets" / "merged" / "Agentic_AI_Experiments_Main_Study002_V1.4.4_270Runs_AnalysisReady.xlsx"
OUTPUT = STUDY_ROOT / "analysis" / "statistical_outputs" / "robustness_sensitivity.xlsx"

DATA_SHEET = "Agentic_AI_Experiments_V1.0"
PROVIDER_COL = "Provider"
WORKFLOW_COL = "workflow_type"
QUALITY_COL = "quality_score"
CONFIDENCE_COL = "confidence"
COST_COL = "cost_usd"
DURATION_COL = "duration_sec"
TOKENS_COL = "total_tokens"
METRICS = [QUALITY_COL, CONFIDENCE_COL, COST_COL, DURATION_COL, TOKENS_COL]
BOOTSTRAP_ITERATIONS = 2000
BOOTSTRAP_SEED = 20260621


def as_float(value: Any) -> float | None:
    if value is None:
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def percentile(values: Sequence[float], p: float) -> float | None:
    vals = sorted(values)
    if not vals:
        return None
    if len(vals) == 1:
        return vals[0]
    rank = (len(vals) - 1) * p
    lower = int(rank)
    upper = min(lower + 1, len(vals) - 1)
    weight = rank - lower
    return vals[lower] * (1 - weight) + vals[upper] * weight


def stddev(values: Sequence[float]) -> float | None:
    vals = list(values)
    if len(vals) < 2:
        return None
    m = mean(vals)
    return sqrt(sum((x - m) ** 2 for x in vals) / (len(vals) - 1))


def bootstrap_mean_ci(values: Sequence[float], iterations: int = BOOTSTRAP_ITERATIONS) -> Tuple[float | None, float | None]:
    vals = list(values)
    if not vals:
        return None, None
    rng = Random(BOOTSTRAP_SEED + len(vals) + int(sum(vals) * 1_000_000) % 1_000_003)
    means: List[float] = []
    n = len(vals)
    for _ in range(iterations):
        sample = [vals[rng.randrange(n)] for _ in range(n)]
        means.append(mean(sample))
    return percentile(means, 0.025), percentile(means, 0.975)


def iqr_bounds(values: Sequence[float]) -> Tuple[float | None, float | None, float | None, float | None]:
    q1 = percentile(values, 0.25)
    q3 = percentile(values, 0.75)
    if q1 is None or q3 is None:
        return None, None, None, None
    iqr = q3 - q1
    return q1, q3, q1 - 1.5 * iqr, q3 + 1.5 * iqr


def trimmed_mean(values: Sequence[float], trim_fraction: float = 0.10) -> float | None:
    vals = sorted(values)
    if not vals:
        return None
    trim = int(len(vals) * trim_fraction)
    if trim == 0 or len(vals) - 2 * trim <= 0:
        return mean(vals)
    return mean(vals[trim:-trim])


def read_rows() -> List[Dict[str, Any]]:
    if not DATASET.exists():
        raise FileNotFoundError(f"Dataset not found: {DATASET}")
    wb = load_workbook(DATASET, read_only=True, data_only=True)
    if DATA_SHEET not in wb.sheetnames:
        raise ValueError(f"Sheet {DATA_SHEET!r} not found. Available: {wb.sheetnames}")
    ws = wb[DATA_SHEET]
    headers = [str(h).strip() if h is not None else "" for h in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
    required = [PROVIDER_COL, WORKFLOW_COL, QUALITY_COL, CONFIDENCE_COL]
    missing = [c for c in required if c not in headers]
    if missing:
        raise ValueError(f"Missing required columns: {missing}")
    rows: List[Dict[str, Any]] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        record = dict(zip(headers, row))
        if record.get(PROVIDER_COL) and record.get(WORKFLOW_COL):
            rows.append(record)
    return rows


def group_rows(rows: List[Dict[str, Any]], group_cols: Tuple[str, ...]) -> Dict[Tuple[Any, ...], List[Dict[str, Any]]]:
    grouped: Dict[Tuple[Any, ...], List[Dict[str, Any]]] = defaultdict(list)
    for row in rows:
        grouped[tuple(row.get(col) for col in group_cols)].append(row)
    return grouped


def metric_summary(rows: List[Dict[str, Any]], group_cols: Tuple[str, ...]) -> List[Dict[str, Any]]:
    records: List[Dict[str, Any]] = []
    grouped = group_rows(rows, group_cols)
    for key in sorted(grouped, key=lambda k: tuple(str(v) for v in k)):
        group = grouped[key]
        for metric in METRICS:
            values = [v for r in group if (v := as_float(r.get(metric))) is not None]
            q1, q3, lower, upper = iqr_bounds(values)
            outlier_count = 0
            non_outlier_values = values
            if lower is not None and upper is not None:
                non_outlier_values = [v for v in values if lower <= v <= upper]
                outlier_count = len(values) - len(non_outlier_values)
            ci_low, ci_high = bootstrap_mean_ci(values) if metric in (QUALITY_COL, CONFIDENCE_COL) else (None, None)
            record: Dict[str, Any] = {col: val for col, val in zip(group_cols, key)}
            record.update(
                {
                    "metric": metric,
                    "N": len(values),
                    "mean": mean(values) if values else None,
                    "median": median(values) if values else None,
                    "q1": q1,
                    "q3": q3,
                    "iqr": (q3 - q1) if q1 is not None and q3 is not None else None,
                    "sd": stddev(values),
                    "min": min(values) if values else None,
                    "max": max(values) if values else None,
                    "trimmed_mean_10pct": trimmed_mean(values),
                    "bootstrap_mean_ci_95_low": ci_low,
                    "bootstrap_mean_ci_95_high": ci_high,
                    "iqr_lower_bound": lower,
                    "iqr_upper_bound": upper,
                    "iqr_outlier_count": outlier_count,
                    "mean_without_iqr_outliers": mean(non_outlier_values) if non_outlier_values else None,
                    "delta_mean_without_outliers": (mean(non_outlier_values) - mean(values)) if values and non_outlier_values else None,
                }
            )
            records.append(record)
    return records


def reviewer_vs_nonreviewer(rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    enriched: List[Dict[str, Any]] = []
    for row in rows:
        copy = dict(row)
        workflow = str(row.get(WORKFLOW_COL))
        copy["reviewer_condition"] = "reviewer" if workflow == "planner_executor_reviewer" else "non_reviewer"
        enriched.append(copy)
    return metric_summary(enriched, ("reviewer_condition",))


def mean_for(rows: List[Dict[str, Any]], metric: str) -> float | None:
    vals = [v for r in rows if (v := as_float(r.get(metric))) is not None]
    return mean(vals) if vals else None


def simple_contrasts(rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    records: List[Dict[str, Any]] = []
    workflows = sorted({str(r.get(WORKFLOW_COL)) for r in rows})
    providers = sorted({str(r.get(PROVIDER_COL)) for r in rows})

    # Workflow contrasts within each provider.
    for provider in providers:
        provider_rows = [r for r in rows if r.get(PROVIDER_COL) == provider]
        for i, left in enumerate(workflows):
            for right in workflows[i + 1 :]:
                left_rows = [r for r in provider_rows if r.get(WORKFLOW_COL) == left]
                right_rows = [r for r in provider_rows if r.get(WORKFLOW_COL) == right]
                for metric in (QUALITY_COL, CONFIDENCE_COL):
                    left_mean = mean_for(left_rows, metric)
                    right_mean = mean_for(right_rows, metric)
                    records.append(
                        {
                            "contrast_scope": "within_provider_workflow_contrast",
                            "fixed_factor": PROVIDER_COL,
                            "fixed_value": provider,
                            "metric": metric,
                            "left_group": left,
                            "right_group": right,
                            "left_N": len(left_rows),
                            "right_N": len(right_rows),
                            "left_mean": left_mean,
                            "right_mean": right_mean,
                            "mean_difference_left_minus_right": (left_mean - right_mean) if left_mean is not None and right_mean is not None else None,
                        }
                    )

    # Provider contrasts within each workflow.
    for workflow in workflows:
        workflow_rows = [r for r in rows if r.get(WORKFLOW_COL) == workflow]
        for i, left in enumerate(providers):
            for right in providers[i + 1 :]:
                left_rows = [r for r in workflow_rows if r.get(PROVIDER_COL) == left]
                right_rows = [r for r in workflow_rows if r.get(PROVIDER_COL) == right]
                for metric in (QUALITY_COL, CONFIDENCE_COL):
                    left_mean = mean_for(left_rows, metric)
                    right_mean = mean_for(right_rows, metric)
                    records.append(
                        {
                            "contrast_scope": "within_workflow_provider_contrast",
                            "fixed_factor": WORKFLOW_COL,
                            "fixed_value": workflow,
                            "metric": metric,
                            "left_group": left,
                            "right_group": right,
                            "left_N": len(left_rows),
                            "right_N": len(right_rows),
                            "left_mean": left_mean,
                            "right_mean": right_mean,
                            "mean_difference_left_minus_right": (left_mean - right_mean) if left_mean is not None and right_mean is not None else None,
                        }
                    )
    return records


def overview(rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    return [
        {"Item": "Dataset", "Value": DATASET.name},
        {"Item": "Total observations", "Value": len(rows)},
        {"Item": "Providers", "Value": len({r.get(PROVIDER_COL) for r in rows})},
        {"Item": "Workflows", "Value": len({r.get(WORKFLOW_COL) for r in rows})},
        {"Item": "Bootstrap iterations", "Value": BOOTSTRAP_ITERATIONS},
        {"Item": "Bootstrap seed", "Value": BOOTSTRAP_SEED},
        {"Item": "Outlier rule", "Value": "1.5 × IQR within each group/metric"},
        {"Item": "Interpretation", "Value": "Robustness outputs are descriptive sensitivity checks for an operational benchmark; they do not replace the planned inferential analyses."},
    ]


def autosize(ws) -> None:
    for col_idx, column_cells in enumerate(ws.columns, start=1):
        max_len = 0
        for cell in column_cells:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 12), 72)


def write_sheet(wb: Workbook, name: str, records: List[Dict[str, Any]]) -> None:
    ws = wb.create_sheet(name)
    if not records:
        ws.append(["No records"])
        return
    headers = list(records[0].keys())
    ws.append(headers)
    for record in records:
        ws.append([record.get(h) for h in headers])
    fill = PatternFill(start_color="375623", end_color="375623", fill_type="solid")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = fill
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            if isinstance(cell.value, float):
                cell.number_format = "0.0000"
    ws.freeze_panes = "A2"
    autosize(ws)


def run() -> None:
    rows = read_rows()
    wb = Workbook()
    default_sheet = wb.active
    if default_sheet is not None:
        wb.remove(default_sheet)

    write_sheet(wb, "Overview", overview(rows))
    write_sheet(wb, "By_Workflow", metric_summary(rows, (WORKFLOW_COL,)))
    write_sheet(wb, "By_Provider", metric_summary(rows, (PROVIDER_COL,)))
    write_sheet(wb, "Provider_Workflow", metric_summary(rows, (PROVIDER_COL, WORKFLOW_COL)))
    write_sheet(wb, "Reviewer_vs_NonReviewer", reviewer_vs_nonreviewer(rows))
    write_sheet(wb, "Simple_Contrasts", simple_contrasts(rows))

    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT)

    print("Robustness/sensitivity analysis completed.")
    print(f"Input rows: {len(rows)}")
    print(f"Output saved: {OUTPUT}")
    for rec in metric_summary(rows, (WORKFLOW_COL,)):
        if rec["metric"] == QUALITY_COL:
            print(
                f"- {rec[WORKFLOW_COL]} quality_score: mean={rec['mean']:.4f}, "
                f"median={rec['median']:.4f}, IQR={rec['iqr']:.4f}, "
                f"95% bootstrap CI=[{rec['bootstrap_mean_ci_95_low']:.4f}, {rec['bootstrap_mean_ci_95_high']:.4f}]"
            )


if __name__ == "__main__":
    run()
