#!/usr/bin/env python3
"""Study 002 measurement-validity audit.

This script audits the relationship between `quality_score` and `confidence`
in the Study 002 AnalysisReady dataset. It is intentionally dependency-light
and uses only openpyxl plus Python's standard library so it can run in minimal
review environments.

Outputs:
  analysis/statistical_outputs/measurement_validity_audit.xlsx

The audit supports the manuscript's public-safe framing of quality_score as an
operational quality proxy rather than an independent human judgment.

Workflow context:
  - Basic Agent and Planner-Executor workflows do not include an independent
    reviewer stage. In V1.4.4 workflow parsing logic, quality_score falls back
    to confidence when no parsed quality_score is emitted.
  - Planner-Executor-Reviewer workflows include a reviewer stage that emits a
    review-derived quality_score, so quality_score and confidence can diverge.
"""

from __future__ import annotations

from collections import defaultdict
from pathlib import Path
from statistics import mean
from typing import Any, Dict, Iterable, List, Tuple

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
STUDY_ROOT = ROOT.parent
DATASET = STUDY_ROOT / "datasets" / "merged" / "Agentic_AI_Experiments_Main_Study002_V1.4.4_270Runs_AnalysisReady.xlsx"
OUTPUT = STUDY_ROOT / "analysis" / "statistical_outputs" / "measurement_validity_audit.xlsx"

DATA_SHEET = "Agentic_AI_Experiments_V1.0"
PROVIDER_COL = "Provider"
WORKFLOW_COL = "workflow_type"
QUALITY_COL = "quality_score"
CONFIDENCE_COL = "confidence"
TASK_CATEGORY_COL = "task_category_final"
TASK_DIFFICULTY_COL = "task_difficulty_final"
TOLERANCE = 1e-12


def as_float(value: Any) -> float | None:
    if value is None:
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def is_equal(a: Any, b: Any, tolerance: float = TOLERANCE) -> bool:
    af = as_float(a)
    bf = as_float(b)
    if af is None or bf is None:
        return False
    return abs(af - bf) <= tolerance


def safe_mean(values: Iterable[float]) -> float | None:
    vals = list(values)
    return mean(vals) if vals else None


def pct(part: int, whole: int) -> float | None:
    return part / whole if whole else None


def read_rows() -> List[Dict[str, Any]]:
    if not DATASET.exists():
        raise FileNotFoundError(f"Dataset not found: {DATASET}")

    wb = load_workbook(DATASET, read_only=True, data_only=True)
    if DATA_SHEET not in wb.sheetnames:
        raise ValueError(f"Sheet {DATA_SHEET!r} not found. Available: {wb.sheetnames}")

    ws = wb[DATA_SHEET]
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    headers = [str(h).strip() if h is not None else "" for h in header_row]
    required = [PROVIDER_COL, WORKFLOW_COL, QUALITY_COL, CONFIDENCE_COL]
    missing = [c for c in required if c not in headers]
    if missing:
        raise ValueError(f"Missing required columns: {missing}")

    rows: List[Dict[str, Any]] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        record = dict(zip(headers, row))
        if any(record.get(c) is not None for c in required):
            rows.append(record)
    return rows


def summarize_group(rows: List[Dict[str, Any]], group_cols: Tuple[str, ...]) -> List[Dict[str, Any]]:
    grouped: Dict[Tuple[Any, ...], List[Dict[str, Any]]] = defaultdict(list)
    for row in rows:
        key = tuple(row.get(col) for col in group_cols)
        grouped[key].append(row)

    output: List[Dict[str, Any]] = []
    for key in sorted(grouped, key=lambda x: tuple(str(v) for v in x)):
        group = grouped[key]
        n = len(group)
        equal_count = sum(is_equal(r.get(QUALITY_COL), r.get(CONFIDENCE_COL)) for r in group)
        diffs: List[float] = []
        abs_diffs: List[float] = []
        q_vals: List[float] = []
        c_vals: List[float] = []
        for r in group:
            q = as_float(r.get(QUALITY_COL))
            c = as_float(r.get(CONFIDENCE_COL))
            if q is not None:
                q_vals.append(q)
            if c is not None:
                c_vals.append(c)
            if q is not None and c is not None:
                d = q - c
                diffs.append(d)
                abs_diffs.append(abs(d))

        row_out: Dict[str, Any] = {col: value for col, value in zip(group_cols, key)}
        row_out.update(
            {
                "N": n,
                "quality_equals_confidence_count": equal_count,
                "quality_equals_confidence_pct": pct(equal_count, n),
                "quality_differs_from_confidence_count": n - equal_count,
                "quality_differs_from_confidence_pct": pct(n - equal_count, n),
                "mean_quality_score": safe_mean(q_vals),
                "mean_confidence": safe_mean(c_vals),
                "mean_quality_minus_confidence": safe_mean(diffs),
                "mean_absolute_difference": safe_mean(abs_diffs),
                "interpretation": interpretation(equal_count, n),
            }
        )
        output.append(row_out)
    return output


def interpretation(equal_count: int, n: int) -> str:
    if n == 0:
        return "No observations."
    ratio = equal_count / n
    if ratio == 1:
        return "Quality proxy is identical to confidence for all observations in this group."
    if ratio >= 0.8:
        return "Quality proxy is mostly identical to confidence in this group."
    if ratio <= 0.2:
        return "Quality proxy is mostly distinct from confidence in this group."
    return "Quality proxy and confidence partially overlap in this group."


def autosize(ws) -> None:
    for col_idx, column_cells in enumerate(ws.columns, start=1):
        max_len = 0
        for cell in column_cells:
            value = cell.value
            if value is not None:
                max_len = max(max_len, len(str(value)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 12), 70)


def write_sheet(wb: Workbook, name: str, records: List[Dict[str, Any]]) -> None:
    ws = wb.create_sheet(name)
    if not records:
        ws.append(["No records"])
        return

    headers = list(records[0].keys())
    ws.append(headers)
    for record in records:
        ws.append([record.get(h) for h in headers])

    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            if isinstance(cell.value, float):
                if "pct" in str(ws.cell(row=1, column=cell.column).value).lower():
                    cell.number_format = "0.0%"
                else:
                    cell.number_format = "0.0000"
    autosize(ws)
    ws.freeze_panes = "A2"


def build_overview(rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    total = len(rows)
    equal_count = sum(is_equal(r.get(QUALITY_COL), r.get(CONFIDENCE_COL)) for r in rows)
    distinct_count = total - equal_count
    return [
        {"Item": "Dataset", "Value": DATASET.name},
        {"Item": "Total observations", "Value": total},
        {"Item": "Providers", "Value": len({r.get(PROVIDER_COL) for r in rows})},
        {"Item": "Workflows", "Value": len({r.get(WORKFLOW_COL) for r in rows})},
        {"Item": "Quality equals confidence count", "Value": equal_count},
        {"Item": "Quality equals confidence percent", "Value": pct(equal_count, total)},
        {"Item": "Quality differs from confidence count", "Value": distinct_count},
        {"Item": "Quality differs from confidence percent", "Value": pct(distinct_count, total)},
        {"Item": "Purpose", "Value": "Audit whether the operational quality proxy is independent from confidence across Study 002 configurations."},
        {"Item": "Workflow design note", "Value": "Basic Agent and Planner-Executor have no reviewer stage; their V1.4.4 parsing logic uses confidence as the quality_score fallback when no parsed quality_score is emitted."},
        {"Item": "Reviewer workflow note", "Value": "Planner-Executor-Reviewer includes a reviewer stage that emits a review-derived quality_score, allowing quality_score and confidence to diverge."},
        {"Item": "Interpretation note", "Value": "High equality rates indicate confidence-aligned operational scoring rather than independent human quality judgment."},
    ]


def run() -> None:
    rows = read_rows()
    wb = Workbook()
    default_sheet = wb.active
    if default_sheet is not None:
        wb.remove(default_sheet)

    write_sheet(wb, "Overview", build_overview(rows))
    write_sheet(wb, "By_Workflow", summarize_group(rows, (WORKFLOW_COL,)))
    write_sheet(wb, "By_Provider", summarize_group(rows, (PROVIDER_COL,)))
    write_sheet(wb, "Provider_Workflow", summarize_group(rows, (PROVIDER_COL, WORKFLOW_COL)))

    if TASK_CATEGORY_COL in rows[0]:
        write_sheet(wb, "By_Task_Category", summarize_group(rows, (TASK_CATEGORY_COL,)))
    if TASK_DIFFICULTY_COL in rows[0]:
        write_sheet(wb, "By_Task_Difficulty", summarize_group(rows, (TASK_DIFFICULTY_COL,)))

    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT)

    by_workflow = summarize_group(rows, (WORKFLOW_COL,))
    print("Measurement-validity audit completed.")
    print(f"Input rows: {len(rows)}")
    print(f"Output saved: {OUTPUT}")
    print("Workflow equality summary:")
    for r in by_workflow:
        print(
            f"- {r[WORKFLOW_COL]}: "
            f"{r['quality_equals_confidence_count']}/{r['N']} "
            f"({r['quality_equals_confidence_pct']:.1%}) equal"
        )


if __name__ == "__main__":
    run()
